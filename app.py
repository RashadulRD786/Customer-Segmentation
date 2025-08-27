# app.py
from flask import Flask, request, jsonify, render_template, send_file, session
import pandas as pd
import numpy as np
import datetime as dt
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from sklearn.mixture import GaussianMixture
import io
import openpyxl
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
import plotly.express as px
import uuid # For generating session IDs

app = Flask(__name__)
# A secret key is required to use Flask sessions. For production, use a strong, random key.
app.secret_key = 'your_super_secret_key_here_replace_me_with_a_long_random_string_for_security' # !!! CHANGE THIS IN PRODUCTION !!!

# --- Global variable to store processed data temporarily by session ID ---
# IMPORTANT: For production, multi-user applications, this needs a more robust
# session management or database integration (e.g., Flask-Session, Redis, or a proper DB).
# For this demonstration, it allows us to pass data between /analyze and /export_clusters.
processed_data_store = {} # Stores {session_id: {'rfm_df': ..., 'cluster_summary': ..., 'auto_cluster_names': ..., 'original_transaction_data_by_customer': ...}}

# --- Helper Functions (Adapted from Jupyter Notebook Cells) ---

def clean_data(df_raw):
    # Logic from Cell 4, enhanced for robustness
    df_cleaned = df_raw.copy()
    
    # Step 1: Standardize column names early and robustly
    # Strip whitespace from column names
    df_cleaned.columns = [col.strip() for col in df_cleaned.columns]

    # Map original potential names to standardized names
    standard_col_mapping = {
        'customer id': 'Customer_ID',
        'customerid': 'Customer_ID',
        'invoice date': 'Invoice_Date',
        'invoicedate': 'Invoice_Date',
        'invoiceno': 'Invoice',
        'invoice': 'Invoice',
        'price': 'Price',
        'unitprice': 'Price',
        'quantity': 'Quantity'
        # 'description': 'Description', # Added for comprehensive original headers in export
        # 'country': 'Country' # Added for comprehensive original headers in export
    }
    
    cleaned_name_map = {}
    for col in df_cleaned.columns:
        matched_standard_name = next(
            (standard_name for variant, standard_name in standard_col_mapping.items() if col.lower() == variant),
            None
        )
        if matched_standard_name:
            if col != matched_standard_name:
                 cleaned_name_map[col] = matched_standard_name
        else:
            cleaned_name_map[col] = col

    if cleaned_name_map:
        final_rename_dict = {actual: target for actual, target in cleaned_name_map.items() if actual != target}
        if final_rename_dict:
            df_cleaned.rename(columns=final_rename_dict, inplace=True)
            if len(set(df_cleaned.columns)) != len(df_cleaned.columns):
                 raise ValueError("Duplicate standardized column names detected after renaming (e.g., 'Customer ID' and 'customerid' both mapped to 'Customer_ID'). Please ensure unique original column names.")

    # Step 2: Define required columns using the STANDARDIZED names
    standardized_required_cols = ['Customer_ID', 'Invoice', 'Quantity', 'Price', 'Invoice_Date']

    # Step 3: Check for missing standardized columns
    missing_cols_after_standardization = [col for col in standardized_required_cols if col not in df_cleaned.columns]
    if missing_cols_after_standardization:
        raise ValueError(f"Missing essential columns after standardization: {', '.join(missing_cols_after_standardization)}. Please ensure your Excel file contains columns for Customer ID, Invoice, Quantity, Price, and Invoice Date (case and space insensitive).")

    # Now proceed with cleaning using standardized names
    df_cleaned['Customer_ID'] = pd.to_numeric(df_cleaned['Customer_ID'], errors='coerce')
    df_cleaned.dropna(subset=['Customer_ID'], inplace=True)
    df_cleaned['Customer_ID'] = df_cleaned['Customer_ID'].astype(int)

    df_cleaned = df_cleaned[~df_cleaned['Invoice'].astype(str).str.contains('C', na=False)]
    
    df_cleaned['Quantity'] = pd.to_numeric(df_cleaned['Quantity'], errors='coerce')
    df_cleaned['Price'] = pd.to_numeric(df_cleaned['Price'], errors='coerce')
    df_cleaned.dropna(subset=['Quantity', 'Price'], inplace=True)
    df_cleaned = df_cleaned[(df_cleaned['Quantity'] > 0) & (df_cleaned['Price'] > 0)]

    df_cleaned['Invoice_Date'] = pd.to_datetime(df_cleaned['Invoice_Date'], errors='coerce')
    df_cleaned = df_cleaned.dropna(subset=['Invoice_Date'])

    if df_cleaned.empty:
        raise ValueError("No valid transactions remaining after cleaning. Please check your data for quality (e.g., all quantities/prices are zero, or dates are invalid).")

    df_cleaned['TotalPrice'] = df_cleaned['Quantity'] * df_cleaned['Price']
    return df_cleaned

def calculate_rfm(df_cleaned):
    # Logic from Cell 5
    snapshot_date = df_cleaned['Invoice_Date'].max() + dt.timedelta(days=1)
    rfm_df = df_cleaned.groupby('Customer_ID').agg(
        Recency=('Invoice_Date', lambda date: (snapshot_date - date.max()).days),
        Frequency=('Invoice', 'nunique'),
        Monetary=('TotalPrice', 'sum')
    )
    return rfm_df

def scale_features(rfm_df):
    # Logic from Cell 6
    X = rfm_df[['Recency', 'Frequency', 'Monetary']]
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    X_scaled_df = pd.DataFrame(X_scaled, columns=X.columns, index=X.index)
    return X_scaled_df

def perform_gmm_clustering(X_scaled_df, n_clusters):
    # Logic from Cell 10
    if len(X_scaled_df) < n_clusters:
        raise ValueError(f"Cannot perform GMM clustering with {n_clusters} clusters. Only {len(X_scaled_df)} valid customer records found after preprocessing.")
    
    gmm = GaussianMixture(n_components=n_clusters, covariance_type='full', random_state=42)
    cluster_labels = gmm.fit_predict(X_scaled_df)
    return cluster_labels

def profile_clusters(rfm_df, cluster_labels):
    # Logic from Cell 11
    rfm_df_with_clusters = rfm_df.copy()
    rfm_df_with_clusters['Cluster'] = cluster_labels
    
    cluster_summary = rfm_df_with_clusters.groupby('Cluster').agg(
        Recency_Mean=('Recency', 'mean'),
        Recency_Median=('Recency', 'median'),
        Frequency_Mean=('Frequency', 'mean'),
        Frequency_Median=('Frequency', 'median'),
        Monetary_Mean=('Monetary', 'mean'),
        Monetary_Median=('Monetary', 'median'),
        Customer_Count=('Recency', 'count')
    ).round(2)
    cluster_summary['Customer_Percentage'] = (cluster_summary['Customer_Count'] / cluster_summary['Customer_Count'].sum() * 100).round(2)
    
    # --- FIX FOR UNDEFINED CLUSTER ID ---
    # Convert the index (Cluster ID) into a regular column named 'cluster'
    cluster_summary = cluster_summary.reset_index()
    # It usually names it 'Cluster' (capital C) after reset_index(). Rename to lowercase 'cluster'
    if 'Cluster' in cluster_summary.columns:
        cluster_summary.rename(columns={'Cluster': 'cluster'}, inplace=True)

    return cluster_summary

# Your precise interpret_rfm_cluster function from your notebook's Cell 12(new)
def interpret_rfm_cluster(recency_scaled, frequency_scaled, monetary_scaled, thresholds):
    """
    Interprets a cluster based on its scaled RFM values and assigns a name,
    using dynamically provided thresholds.
    Recency_scaled is assumed to be inverted (higher value = more recent, range 0-1).
    """
    R_EXTREME_RECENT_MIN = thresholds['Recency_EXTREME_RECENT_MIN']
    R_VERY_HIGH_MIN = thresholds['Recency_VERY_HIGH_MIN']
    R_MODERATELY_RECENT_MIN = thresholds['Recency_MODERATELY_RECENT_MIN']
    R_INACTIVE_BORDER_MIN = thresholds['Recency_INACTIVE_BORDER_MIN']

    F_EXTREME_HIGH_MIN = thresholds['Frequency_EXTREME_HIGH_MIN']
    F_HIGH_MIN = thresholds['Frequency_HIGH_MIN']
    F_MEDIUM_HIGH_MIN = thresholds['Frequency_MEDIUM_HIGH_MIN']
    F_MEDIUM_LOW_MIN = thresholds['Frequency_MEDIUM_LOW_MIN']
    F_VERY_LOW_MIN = thresholds['Frequency_VERY_LOW_MIN']

    M_EXTREME_HIGH_MIN = thresholds['Monetary_EXTREME_HIGH_MIN']
    M_HIGH_MIN = thresholds['Monetary_HIGH_MIN']
    M_MEDIUM_HIGH_MIN = thresholds['Monetary_MEDIUM_HIGH_MIN']
    M_MEDIUM_LOW_MIN = thresholds['Monetary_MEDIUM_LOW_MIN']
    M_VERY_LOW_MIN = thresholds['Monetary_VERY_LOW_MIN']


    r_level = ""
    if recency_scaled >= R_EXTREME_RECENT_MIN : r_level = "Extreme Recent"
    elif recency_scaled >= R_VERY_HIGH_MIN and recency_scaled < R_EXTREME_RECENT_MIN: r_level = "Very High Recent"
    elif recency_scaled >= R_MODERATELY_RECENT_MIN and recency_scaled < R_VERY_HIGH_MIN: r_level = "Moderately Recent"
    elif recency_scaled >= R_INACTIVE_BORDER_MIN and recency_scaled < R_MODERATELY_RECENT_MIN: r_level = "Inactive Border"
    else: r_level = "Extremely Inactive"

    f_level = ""
    if frequency_scaled >= F_EXTREME_HIGH_MIN: f_level = "Extreme High F"
    elif frequency_scaled >= F_HIGH_MIN : f_level = "High F"
    elif frequency_scaled >= F_MEDIUM_HIGH_MIN: f_level = "Medium High F"
    elif frequency_scaled >= F_MEDIUM_LOW_MIN: f_level = "Medium Low F"
    elif frequency_scaled >= F_VERY_LOW_MIN: f_level = "Very Low F"
    else: f_level = "Extremely Low F"

    m_level = ""
    if monetary_scaled >= M_EXTREME_HIGH_MIN: m_level = "Extreme High M"
    elif monetary_scaled >= M_HIGH_MIN: m_level = "High M"
    elif monetary_scaled >= M_MEDIUM_HIGH_MIN: m_level = "Medium High M"
    elif monetary_scaled >= M_MEDIUM_LOW_MIN: m_level = "Medium Low M"
    elif monetary_scaled >= M_VERY_LOW_MIN: m_level = "Very Low M"
    else: m_level = "Extremely Low M"


    # --- Rule-Based Naming Logic (ORDER IS CRUCIAL) ---
    if (recency_scaled < 1e-15 and frequency_scaled < 1e-15 and monetary_scaled < 1e-15):
        return "Lost Customer (Absolute zero)"
    if r_level == "Extreme Recent" and f_level == "Extreme High F" and m_level == "Extreme High M":
        return "Champions"
    if r_level == "Very High Recent" and f_level == "High F" and m_level == "Medium High M": # C6: R=0.997, F=0.494, M=0.070
        return "Elite loyal customers"
    if r_level == "Very High Recent" and f_level == "High F" and m_level == "High M": # C4: R=0.969, F=0.35, M=0.11
        return "High value loyalists"
    if r_level == "Very High Recent" and f_level == "Medium High F" and m_level == "Medium Low M": # C7: R=0.088, F=0.11, M=0.014
        return "Promising loyalists-Active mid tier"
    if r_level == "Very High Recent" and f_level == "Medium Low F" and m_level == "Very Low M": # C1: R=0.093, F=0.058, M=0.005
        return "New Customers (Potential)"
    if r_level == "Very High Recent" and f_level == "Very Low F" and m_level == "Very Low M": # C8: R=0.090, F=0.016, M=0.001
        return "New customers (Low engagement)"
    if r_level == "Moderately Recent" and f_level == "Medium High F" and m_level == "Medium High M": # C3: R=0.196, F=0.098, M=0.023
        return "At risk"
    if r_level == "Moderately Recent" and f_level == "Medium Low F" and m_level == "Very Low M": # C0: R=0.236, F=0.026, M=0.002
        return "Dormant (inactive)"
    if r_level == "Inactive Border" and f_level == "Very Low F" and m_level == "Very Low M": # C5: R=0.00000005, F=0.008, M=0.0004
        if not (recency_scaled < 1e-15 and frequency_scaled < 1e-15 and monetary_scaled < 1e-15):
            return "Nearly Lost"
    return "Other Segment - Undefined"

def calculate_dynamic_thresholds(cluster_means_for_naming_scaled):
    # Logic from Cell 12(new) - Dynamic Thresholds
    dynamic_thresholds = {}
    if not cluster_means_for_naming_scaled.empty:
        # Loop over the expected base RFM names for the scaled DataFrame
        for col_base in ['Recency', 'Frequency', 'Monetary']: # <-- CORRECTED: Use base names here
            
            # This check is now redundant because `cluster_means_for_naming_scaled`
            # is explicitly created with these column names in `run_full_analysis`.
            # Leaving it for extreme robustness, but `col_name` is just `col_base` now.
            if col_base not in cluster_means_for_naming_scaled.columns:
                raise KeyError(f"Expected column '{col_base}' not found in cluster_means_for_naming_scaled DataFrame. Check how it's being constructed.")

            values = cluster_means_for_naming_scaled[col_base] # <-- Access directly using col_base
            
            if values.min() == values.max():
                if col_base == 'Recency':
                    dynamic_thresholds[col_base + '_EXTREME_RECENT_MIN'] = values.iloc[0] 
                    dynamic_thresholds[col_base + '_VERY_HIGH_MIN'] = values.iloc[0]
                    dynamic_thresholds[col_base + '_MODERATELY_RECENT_MIN'] = values.iloc[0]
                    dynamic_thresholds[col_base + '_INACTIVE_BORDER_MIN'] = values.iloc[0]
                elif col_base == 'Frequency':
                    dynamic_thresholds[col_base + '_EXTREME_HIGH_MIN'] = values.iloc[0]
                    dynamic_thresholds[col_base + '_HIGH_MIN'] = values.iloc[0]
                    dynamic_thresholds[col_base + '_MEDIUM_HIGH_MIN'] = values.iloc[0]
                    dynamic_thresholds[col_base + '_MEDIUM_LOW_MIN'] = values.iloc[0]
                    dynamic_thresholds[col_base + '_VERY_LOW_MIN'] = values.iloc[0]
                elif col_base == 'Monetary':
                    dynamic_thresholds[col_base + '_EXTREME_HIGH_MIN'] = values.iloc[0]
                    dynamic_thresholds[col_base + '_HIGH_MIN'] = values.iloc[0]
                    dynamic_thresholds[col_base + '_MEDIUM_HIGH_MIN'] = values.iloc[0]
                    dynamic_thresholds[col_base + '_MEDIUM_LOW_MIN'] = values.iloc[0]
                    dynamic_thresholds[col_base + '_VERY_LOW_MIN'] = values.iloc[0]
            else:
                if col_base == 'Recency':
                    dynamic_thresholds[col_base + '_EXTREME_RECENT_MIN'] = values.quantile(0.97)
                    dynamic_thresholds[col_base + '_VERY_HIGH_MIN'] = values.quantile(0.4)
                    dynamic_thresholds[col_base + '_MODERATELY_RECENT_MIN'] = values.quantile(0.12)
                    dynamic_thresholds[col_base + '_INACTIVE_BORDER_MIN'] = values.quantile(0.04)
                elif col_base == 'Frequency':
                    dynamic_thresholds[col_base + '_EXTREME_HIGH_MIN'] = values.quantile(0.99)
                    dynamic_thresholds[col_base + '_HIGH_MIN'] = values.quantile(0.75)
                    dynamic_thresholds[col_base + '_MEDIUM_HIGH_MIN'] = values.quantile(0.50)
                    dynamic_thresholds[col_base + '_MEDIUM_LOW_MIN'] = values.quantile(0.3)
                    dynamic_thresholds[col_base + '_VERY_LOW_MIN'] = values.quantile(0.005)
                elif col_base == 'Monetary':
                    dynamic_thresholds[col_base + '_EXTREME_HIGH_MIN'] = values.quantile(0.99)
                    dynamic_thresholds[col_base + '_HIGH_MIN'] = values.quantile(0.85)
                    dynamic_thresholds[col_base + '_MEDIUM_HIGH_MIN'] = values.quantile(0.6)
                    dynamic_thresholds[col_base + '_MEDIUM_LOW_MIN'] = values.quantile(0.55)
                    dynamic_thresholds[col_base + '_VERY_LOW_MIN'] = values.quantile(0.0009)
    return dynamic_thresholds

def assign_cluster_names(cluster_means_for_naming_scaled, dynamic_thresholds):
    # Logic from Cell 12(new) - Application of Interpretation Function
    auto_cluster_names = {}
    name_counts = {}
    if not cluster_means_for_naming_scaled.empty:
        for cluster_id, row in cluster_means_for_naming_scaled.iterrows():
            # Access the scaled means using their base names (e.g., 'Recency')
            base_name = interpret_rfm_cluster(
                row['Recency'], # Use Recency (now directly from scaled DF)
                row['Frequency'], # Use Frequency
                row['Monetary'], # Use Monetary
                dynamic_thresholds
            )
            
            if base_name in name_counts:
                name_counts[base_name] += 1
                auto_cluster_names[cluster_id] = f"{base_name} ({name_counts[base_name]})"
            else:
                name_counts[base_name] = 1
                auto_cluster_names[cluster_id] = base_name
    return auto_cluster_names

def generate_treemap_json(cluster_summary_df, auto_cluster_names, n_clusters_actual):
    # Logic from Cell 13
    treemap_df_data = []
    total_customers_for_treemap = cluster_summary_df['Customer_Count'].sum()

    # Identify the exact name of the 'Champions' cluster from auto_cluster_names
    champion_cluster_name_exact = None
    for cid, cname in auto_cluster_names.items():
        if cname.startswith("Champions"): # Catches "Champions", "Champions (2)", etc.
            champion_cluster_name_exact = cname
            break

    SEGMENTS_TO_ALWAYS_SHOW = []
    if champion_cluster_name_exact:
        SEGMENTS_TO_ALWAYS_SHOW.append(champion_cluster_name_exact)
    SEGMENTS_TO_ALWAYS_SHOW.extend(["Elite loyal customers", "High value loyalists"])

    MIN_PERCENTAGE_FOR_TREEMAP_LABEL = 0.5

    small_clusters_count = 0
    small_clusters_percentage = 0.0
    small_clusters_names = []

    for cluster_id, row in cluster_summary_df.iterrows():
        # Ensure cluster_id is correctly fetched from the 'cluster' column, not the index anymore
        current_cluster_id = row['cluster'] 
        cluster_name = auto_cluster_names.get(current_cluster_id, f"Cluster {current_cluster_id}")
        
        if cluster_name in SEGMENTS_TO_ALWAYS_SHOW:
            treemap_df_data.append({
                'ClusterID': current_cluster_id, # Use current_cluster_id
                'ClusterName': cluster_name,
                'CustomerCount': row['Customer_Count'],
                'Percentage': row['Customer_Percentage'],
                'Recency_Mean': row['Recency_Mean'],
                'Frequency_Mean': row['Frequency_Mean'],
                'Monetary_Mean': row['Monetary_Mean'],
                'TotalCustomers': total_customers_for_treemap,
                'GroupedNames': ''
            })
        elif row['Customer_Percentage'] >= MIN_PERCENTAGE_FOR_TREEMAP_LABEL:
            treemap_df_data.append({
                'ClusterID': current_cluster_id, # Use current_cluster_id
                'ClusterName': cluster_name,
                'CustomerCount': row['Customer_Count'],
                'Percentage': row['Customer_Percentage'],
                'Recency_Mean': row['Recency_Mean'],
                'Frequency_Mean': row['Frequency_Mean'],
                'Monetary_Mean': row['Monetary_Mean'],
                'TotalCustomers': total_customers_for_treemap,
                'GroupedNames': ''
            })
        else:
            small_clusters_count += row['Customer_Count']
            small_clusters_percentage += row['Customer_Percentage']
            small_clusters_names.append(cluster_name)

    if small_clusters_count > 0:
        # Filter by names because original cluster_summary_df index is no longer directly the cluster_id
        clusters_in_other = cluster_summary_df[cluster_summary_df['Auto_Cluster_Name'].isin(small_clusters_names)]
        
        other_recency_mean = clusters_in_other['Recency_Mean'].mean() if not clusters_in_other.empty else 0
        other_frequency_mean = clusters_in_other['Frequency_Mean'].mean() if not clusters_in_other.empty else 0
        other_monetary_mean = clusters_in_other['Monetary_Mean'].mean() if not clusters_in_other.empty else 0
        
        treemap_df_data.append({
            'ClusterID': -1,
            'ClusterName': f"Other Segments",
            'CustomerCount': small_clusters_count,
            'Percentage': small_clusters_percentage,
            'Recency_Mean': other_recency_mean,
            'Frequency_Mean': other_frequency_mean,
            'Monetary_Mean': other_monetary_mean,
            'TotalCustomers': total_customers_for_treemap,
            'GroupedNames': "<br>Included Clusters:<br>" + "<br>".join(small_clusters_names)
        })

    treemap_df = pd.DataFrame(treemap_df_data)

    if treemap_df.empty:
        return None # No data to plot

    colors_palette = px.colors.qualitative.D3
    cluster_color_map = {
        cluster_id: colors_palette[cluster_id % len(colors_palette)]
        for cluster_id in sorted(treemap_df['ClusterID'].unique()) if cluster_id != -1
    }
    cluster_color_map[-1] = '#A9A9A9' # DarkGrey for 'Other'
    segment_colors = [cluster_color_map[cid] for cid in treemap_df['ClusterID']]

    fig = go.Figure(go.Treemap(
        labels=treemap_df['ClusterName'],
        parents=['All Customers'] * len(treemap_df), # All direct children of 'All Customers'
        values=treemap_df['CustomerCount'],
        marker=dict(colors=segment_colors),
        hovertemplate="<b>%{label}</b><br><br>" +
                      "Customers: %{customdata[3]} / %{customdata[5]} (%{customdata[4]:.1f}%)<br>" +
                      "Recency: %{customdata[0]:.0f} days<br>" +
                      "Frequency: %{customdata[1]:.1f} purchases<br>" +
                      "Monetary: $%{customdata[2]:.2f}%{customdata[6]}<extra></extra>",
        customdata=treemap_df[['Recency_Mean', 'Frequency_Mean', 'Monetary_Mean',
                               'CustomerCount', 'Percentage', 'TotalCustomers', 'GroupedNames']].values.tolist(),
        textinfo="label+value"
    ))

    fig.update_layout(
        margin = dict(t=50, l=25, r=25, b=25),
        title_text=f"Interactive Treemap: Customer Segments ({n_clusters_actual} Clusters)",
        font=dict(family='Inter, sans-serif')
    )

    return fig.to_json() # Convert Plotly figure to JSON string

# Main Analysis Orchestrator
def run_full_analysis(file_content, n_clusters_requested):
    xls = pd.ExcelFile(io.BytesIO(file_content))
    
    df_2009_2010 = pd.DataFrame()
    df_2010_2011 = pd.DataFrame()

    if 'Year 2009-2010' in xls.sheet_names:
        df_2009_2010 = xls.parse('Year 2009-2010')
    if 'Year 2010-2011' in xls.sheet_names:
        df_2010_2011 = xls.parse('Year 2010-2011')

    if df_2009_2010.empty and df_2010_2011.empty:
        raise ValueError("No data found in 'Year 2009-2010' or 'Year 2010-2011' sheets. Please ensure sheet names are correct and contain data.")
    
    df_combined = pd.concat([df_2009_2010, df_2010_2011], ignore_index=True)

    df_cleaned = clean_data(df_combined)
    
    customer_original_transactions_map = df_cleaned.groupby('Customer_ID').apply(lambda x: x.to_dict(orient='records')).to_dict()

    rfm_df = calculate_rfm(df_cleaned)
    if rfm_df.empty:
        raise ValueError("No RFM data could be calculated after cleaning. Check customer IDs, dates, quantities, and prices in your data.")

    X_scaled_df = scale_features(rfm_df)
    
    if len(X_scaled_df) < n_clusters_requested:
        raise ValueError(f"Not enough distinct customer records ({len(X_scaled_df)}) to create {n_clusters_requested} clusters. Please reduce the number of clusters or provide more data.")
        
    cluster_labels = perform_gmm_clustering(X_scaled_df, n_clusters_requested)
    
    rfm_df_with_clusters = rfm_df.copy()
    rfm_df_with_clusters['Cluster'] = cluster_labels

    cluster_summary_df = profile_clusters(rfm_df_with_clusters, cluster_labels) # This returns DataFrame with 'cluster' column
    
    actual_n_clusters_found = len(cluster_summary_df)
    n_clusters_display = actual_n_clusters_found # Always use the actual count for display

    min_max_scaler_for_naming = MinMaxScaler()
    # Create the DataFrame for scaling with base RFM names, derived from the _Mean columns
    cluster_means_for_naming_raw = cluster_summary_df[['Recency_Mean', 'Frequency_Mean', 'Monetary_Mean']].copy()

    cluster_means_for_naming_scaled_values = min_max_scaler_for_naming.fit_transform(cluster_means_for_naming_raw)
    
    cluster_means_for_naming_scaled = pd.DataFrame(
        cluster_means_for_naming_scaled_values,
        columns=['Recency', 'Frequency', 'Monetary'], # These are the columns `calculate_dynamic_thresholds` expects
        index=cluster_means_for_naming_raw.index # Preserve cluster IDs as index
    )

    cluster_means_for_naming_scaled['Recency'] = 1 - cluster_means_for_naming_scaled['Recency'] 

    dynamic_thresholds = calculate_dynamic_thresholds(cluster_means_for_naming_scaled) # Pass this DataFrame
    auto_cluster_names = assign_cluster_names(cluster_means_for_naming_scaled, dynamic_thresholds) # Pass this DataFrame

    # Update cluster_summary_df with auto-assigned names using its 'cluster' column
    cluster_summary_df['Auto_Cluster_Name'] = cluster_summary_df['cluster'].map(auto_cluster_names)
    
    all_customers_data_for_export = []
    for customer_id, rfm_row in rfm_df_with_clusters.iterrows():
        all_customers_data_for_export.append({
            'Customer_ID': customer_id,
            'RFM_Recency': rfm_row['Recency'],
            'RFM_Frequency': rfm_row['Frequency'],
            'RFM_Monetary': rfm_row['Monetary'],
            'Assigned_Cluster': rfm_row['Cluster'],
            'Original_Transactions': customer_original_transactions_map.get(customer_id, [])
        })

    treemap_json = generate_treemap_json(cluster_summary_df, auto_cluster_names, n_clusters_display)

    return cluster_summary_df.to_dict(orient='records'), auto_cluster_names, treemap_json, all_customers_data_for_export

# --- Flask Routes ---

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/analyze', methods=['POST'])
def analyze():
    current_session_id = str(uuid.uuid4())
    session['analysis_session_id'] = current_session_id

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file part in the request.'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file.'}), 400
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'message': 'Invalid file type. Please upload an .xlsx file.'}), 400

    try:
        n_clusters = int(request.form.get('nClusters', 5))
        if not (2 <= n_clusters <= 10):
            raise ValueError('Number of clusters must be between 2 and 10.')

        file_content = file.read()
        
        cluster_summary_data_list, auto_cluster_names, treemap_json, all_customers_data_for_export = \
            run_full_analysis(file_content, n_clusters)

        processed_data_store[current_session_id] = {
            'cluster_summary': cluster_summary_data_list,
            'auto_cluster_names': auto_cluster_names,
            'treemap_json': treemap_json,
            'all_customers_data_for_export': all_customers_data_for_export
        }
        
        return jsonify({
            'success': True,
            'message': 'Analysis complete! Check results below.',
            'clusterSummary': cluster_summary_data_list,
            'autoClusterNames': auto_cluster_names,
            'treemapJson': treemap_json,
            'nClusters': n_clusters # Frontend still expects nClusters for title, can be adjusted
        })

    except ValueError as ve:
        app.logger.error(f"Analysis input/data error: {ve}", exc_info=True)
        return jsonify({'success': False, 'message': f'Data Error: {str(ve)}'}), 400
    except Exception as e:
        app.logger.error(f"Analysis general error: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Analysis failed: An unexpected error occurred. {str(e)}. Please check server logs for details.'}), 500

@app.route('/export_clusters', methods=['GET'])
def export_clusters():
    current_session_id = session.get('analysis_session_id')
    processed_data = processed_data_store.get(current_session_id)

    if not processed_data:
        return jsonify({'success': False, 'message': 'No analysis results found for this session to export. Please run analysis first.'}), 400

    cluster_summary_data = processed_data['cluster_summary']
    auto_cluster_names = processed_data['auto_cluster_names']
    all_customers_data_for_export = processed_data['all_customers_data_for_export']

    try:
        output = io.BytesIO()
        workbook = openpyxl.Workbook()

        # 1. Summary Sheet
        summary_ws = workbook.active
        summary_ws.title = "Cluster Summary"
        summary_headers = ["Cluster ID", "Auto Cluster Name", "Customer Count", "Customer Percentage", "Recency Mean", "Frequency Mean", "Monetary Mean"]
        summary_ws.append(summary_headers)
        for cluster_data in cluster_summary_data:
            summary_ws.append([
                cluster_data['cluster'],
                cluster_data['Auto_Cluster_Name'],
                cluster_data['Customer_Count'],
                cluster_data['Customer_Percentage'],
                cluster_data['Recency_Mean'],
                cluster_data['Frequency_Mean'],
                cluster_data['Monetary_Mean']
            ])
        
        for column in summary_ws.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            summary_ws.column_dimensions[column_name].width = adjusted_width


        # 2. Individual Cluster Sheets
        sorted_cluster_ids = sorted([c['cluster'] for c in cluster_summary_data])

        all_possible_original_headers = set()
        for customer in all_customers_data_for_export:
            for original_tx in customer['Original_Transactions']:
                all_possible_original_headers.update(original_tx.keys())
        
        known_original_order = ['Invoice', 'StockCode', 'Description', 'Quantity', 'Invoice_Date', 'Price', 'Customer_ID', 'Country', 'TotalPrice']
        original_headers_sorted = [h for h in known_original_order if h in all_possible_original_headers]
        for h in sorted(list(all_possible_original_headers)):
            if h not in original_headers_sorted:
                original_headers_sorted.append(h)

        rfm_cluster_headers = ["RFM_Recency", "RFM_Frequency", "RFM_Monetary", "Assigned_Cluster"]

        for cluster_id in sorted_cluster_ids:
            cluster_full_name = auto_cluster_names.get(cluster_id, f"Cluster {cluster_id}")
            sheet_name = f"C{cluster_id} - {cluster_full_name}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:28] + '...'

            cluster_ws = workbook.create_sheet(title=sheet_name)
            
            customers_in_this_cluster = [
                cust for cust in all_customers_data_for_export if cust['Assigned_Cluster'] == cluster_id
            ]

            full_sheet_headers = ['Customer_ID'] + original_headers_sorted + rfm_cluster_headers
            cluster_ws.append(full_sheet_headers)

            for customer in customers_in_this_cluster:
                customer_id = customer['Customer_ID']
                rfm_recency = customer['RFM_Recency']
                rfm_frequency = customer['RFM_Frequency']
                rfm_monetary = customer['RFM_Monetary'] 
                assigned_cluster = customer['Assigned_Cluster']

                if customer['Original_Transactions']:
                    for original_tx in customer['Original_Transactions']:
                        row_data = [customer_id]
                        for header in original_headers_sorted:
                            row_data.append(original_tx.get(header, ''))
                        row_data.extend([rfm_recency, rfm_frequency, rfm_monetary, assigned_cluster])
                        cluster_ws.append(row_data)
                else:
                    row_data = [customer_id] + [''] * len(original_headers_sorted) + [rfm_recency, rfm_frequency, rfm_monetary, assigned_cluster]
                    cluster_ws.append(row_data)
        
            for column in cluster_ws.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                cluster_ws.column_dimensions[column_name].width = adjusted_width

        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        workbook.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name='customer_segmentation_report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        app.logger.error(f"Export error: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Export failed: {str(e)}. Check server logs for details.'}), 500

if __name__ == '__main__':
    app.run(debug=True)